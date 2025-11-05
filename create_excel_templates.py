"""
엑셀 업로드 양식 템플릿 생성 스크립트
각 파일 타입별 샘플 데이터가 포함된 템플릿 생성
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def create_template_excel(filename, sheet_name, data, required_columns, optional_columns=None):
    """
    엑셀 템플릿 생성
    
    Args:
        filename: 파일명
        sheet_name: 시트명
        data: 샘플 데이터 (DataFrame)
        required_columns: 필수 컬럼 리스트
        optional_columns: 선택 컬럼 리스트
    """
    # 디렉토리 생성
    output_dir = 'excel_templates'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    filepath = os.path.join(output_dir, filename)
    
    # ExcelWriter로 생성
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 스타일 적용
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, column_title in enumerate(data.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 필수 컬럼 표시
            if column_title in required_columns:
                cell.value = f"{column_title} *"
        
        # 컬럼 너비 자동 조정
        for col_num, column_title in enumerate(data.columns, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(
                len(str(column_title)),
                *[len(str(value)) for value in data[column_title].values]
            )
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
        # 설명 시트 추가
        if optional_columns:
            ws_info = workbook.create_sheet("설명")
            ws_info.append(["컬럼명", "타입", "필수여부", "설명"])
            
            all_columns = required_columns + optional_columns
            for col in all_columns:
                is_required = col in required_columns
                ws_info.append([
                    col,
                    "텍스트/숫자",
                    "필수" if is_required else "선택",
                    f"{col} 설명"
                ])
    
    print(f"[OK] {filename} created successfully")

# 1. 그룹/멤버사 연간 학습시간
annual_data = pd.DataFrame({
    '멤버사명': ['SK텔레콤', 'SK텔레콤', 'SK텔레콤', 'SK텔레콤', 
                'SK하이닉스', 'SK하이닉스', 'SK하이닉스', 'SK하이닉스'],
    '연도': [2022, 2023, 2024, 2025, 2022, 2023, 2024, 2025],
    '학습시간': [1000.0, 1150.5, 1200.0, 1250.5, 800.0, 950.0, 1100.0, 1150.0],
    '전년대비변화율': [None, 15.05, 4.30, 4.21, None, 18.75, 15.79, 4.55],
    '상반기_학습시간': [500.0, 550.0, 580.0, 600.0, 400.0, 475.0, 550.0, 575.0],
    '하반기_학습시간': [500.0, 600.5, 620.0, 650.5, 400.0, 475.0, 550.0, 575.0]
})

create_template_excel(
    '01_그룹_멤버사_연간_학습시간.xlsx',
    '연간학습시간',
    annual_data,
    ['멤버사명', '연도', '학습시간'],
    ['전년대비변화율', '상반기_학습시간', '하반기_학습시간']
)

# 2. 멤버사 월별 학습시간
monthly_data = pd.DataFrame({
    '멤버사명': ['SK텔레콤'] * 12,
    '연도': [2025] * 12,
    '월': list(range(1, 13)),
    '학습시간': [95.5, 102.0, 110.5, 105.0, 98.5, 103.0,
                107.5, 112.0, 108.0, 104.5, 99.0, 106.5]
})

create_template_excel(
    '02_멤버사_월별_학습시간.xlsx',
    '월별학습시간',
    monthly_data,
    ['멤버사명', '연도', '월', '학습시간']
)

# 3. 학습 카테고리별 학습시간
category_data = pd.DataFrame({
    '카테고리명': ['AI/DT', '경영철학', '공통직무역량', '리더십', '디지털 트랜스포메이션'],
    '학습시간': [5000.0, 3500.0, 4200.0, 2800.0, 3200.0],
    '학습자수': [250, 180, 200, 140, 160]
})

create_template_excel(
    '03_학습_카테고리별_학습시간.xlsx',
    '카테고리별학습시간',
    category_data,
    ['카테고리명', '학습시간'],
    ['학습자수']
)

# 4. 인기 학습카드
popular_cards_data = pd.DataFrame({
    '학습카드명': ['ChatGPT 활용하기', '데이터 분석 기초', '프롬프트 엔지니어링',
                 'Python 기초', '머신러닝 입문', '디지털 마케팅', '프로젝트 관리',
                 '커뮤니케이션 스킬', '리더십 개발', '경영 전략'],
    '학습자수': [150, 120, 110, 105, 100, 95, 90, 85, 80, 75],
    '총학습시간': [750.0, 600.0, 550.0, 525.0, 500.0, 475.0, 450.0, 425.0, 400.0, 375.0],
    '카테고리': ['AI/DT', 'AI/DT', 'AI/DT', 'AI/DT', 'AI/DT', '마케팅', 
                '프로젝트관리', '공통직무', '리더십', '경영철학']
})

create_template_excel(
    '04_인기_학습카드.xlsx',
    '인기학습카드',
    popular_cards_data,
    ['학습카드명', '학습자수'],
    ['총학습시간', '카테고리']
)

# 5. 검색어 데이터
search_data = pd.DataFrame({
    '검색어': ['AI', '데이터 분석', 'ChatGPT', 'Python', '머신러닝',
             'AI', '데이터 분석', 'ChatGPT', 'Python', '머신러닝'],
    '연도': [2024, 2024, 2024, 2024, 2024,
            2025, 2025, 2025, 2025, 2025],
    '검색횟수': [1500, 1200, 1100, 1000, 950,
                1800, 1400, 1300, 1200, 1100]
})

create_template_excel(
    '05_검색어_데이터.xlsx',
    '검색어',
    search_data,
    ['검색어', '연도'],
    ['검색횟수']
)

# 6. 주요 영역 인증/이수 현황표
area_data = pd.DataFrame({
    '멤버사명': ['SK텔레콤', 'SK텔레콤', 'SK텔레콤',
                'SK하이닉스', 'SK하이닉스', 'SK하이닉스'],
    '영역명': ['경영철학', 'AI/DT', '공통직무역량',
              '경영철학', 'AI/DT', '공통직무역량'],
    '이수인원': [250, 300, 280, 200, 250, 230],
    '인증인원': [200, 250, 220, 160, 200, 180],
    '이수율': [80.0, 83.3, 78.6, 80.0, 80.0, 78.3]
})

create_template_excel(
    '06_주요_영역_인증_이수_현황.xlsx',
    '영역별현황',
    area_data,
    ['영역명', '이수인원'],
    ['멤버사명', '인증인원', '이수율']
)

# 7. 개인별 학습시간 raw data
individual_data = pd.DataFrame({
    '개인ID': ['EMP001', 'EMP002', 'EMP003', 'EMP004', 'EMP005'],
    '이름': ['홍길동', '김철수', '이영희', '박민수', '정수진'],
    '조직': ['경영지원팀', 'IT팀', '마케팅팀', '영업팀', 'HR팀'],
    '직책': ['임원', '팀장', '구성원', '팀장', '구성원'],
    '학습시간': [125.5, 98.0, 75.5, 110.0, 85.0]
})

create_template_excel(
    '07_개인별_학습시간_raw.xlsx',
    '개인별학습시간',
    individual_data,
    ['개인ID', '학습시간'],
    ['이름', '조직', '직책']
)

# 8. 카드별 학습시간 raw data
card_data = pd.DataFrame({
    '학습카드ID': ['CARD001', 'CARD002', 'CARD003', 'CARD004', 'CARD005'],
    '학습카드명': ['ChatGPT 활용하기', '데이터 분석 기초', '프롬프트 엔지니어링',
                 'Python 기초', '머신러닝 입문'],
    '생성정보': ['2024-01-15', '2024-02-10', '2024-03-05',
                '2024-01-20', '2024-02-15'],
    '이수인원': [150, 120, 110, 105, 100],
    '현재학습인원': [20, 15, 12, 18, 10],
    '학습시간': [750.0, 600.0, 550.0, 525.0, 500.0]
})

create_template_excel(
    '08_카드별_학습시간_raw.xlsx',
    '카드별학습시간',
    card_data,
    ['학습카드ID', '학습카드명'],
    ['생성정보', '이수인원', '현재학습인원', '학습시간']
)

# 9. Badge별 학습시간 raw data
badge_data = pd.DataFrame({
    'BadgeID': ['BADGE001', 'BADGE002', 'BADGE003', 'BADGE004', 'BADGE005'],
    'Badge명': ['AI 마스터', '데이터 분석가', '프롬프트 전문가',
               'Python 개발자', '머신러닝 엔지니어'],
    '기본정보': ['AI 분야 마스터 인증', '데이터 분석 전문가 인증',
                '프롬프트 엔지니어링 전문가', 'Python 개발 역량 인증',
                '머신러닝 전문가 인증'],
    '취득인원': [50, 40, 35, 45, 30],
    '도전중인원': [100, 80, 70, 90, 60],
    '학습시간': [5000.0, 4000.0, 3500.0, 4500.0, 3000.0]
})

create_template_excel(
    '09_Badge별_학습시간_raw.xlsx',
    'Badge별학습시간',
    badge_data,
    ['BadgeID', 'Badge명'],
    ['기본정보', '취득인원', '도전중인원', '학습시간']
)

# 10. 개인별 학습 전체 raw data (변화군 분석용)
individual_full_data = pd.DataFrame({
    '개인ID': ['EMP001', 'EMP002', 'EMP003', 'EMP004', 'EMP005'],
    '이름': ['홍길동', '김철수', '이영희', '박민수', '정수진'],
    '조직': ['경영지원팀', 'IT팀', '마케팅팀', '영업팀', 'HR팀'],
    '직책': ['임원', '팀장', '구성원', '팀장', '구성원'],
    '24년학습시간': [100.0, 80.0, 60.0, 90.0, 70.0],
    '25년학습시간': [125.5, 98.0, 75.5, 110.0, 85.0],
    '학습카드명': ['ChatGPT 활용하기', '데이터 분석 기초', '프롬프트 엔지니어링',
                 'Python 기초', '머신러닝 입문'],
    '학습일시': ['2025-01-15 10:30:00', '2025-01-16 14:20:00',
                '2025-01-17 09:15:00', '2025-01-18 11:00:00',
                '2025-01-19 15:30:00']
})

create_template_excel(
    '10_개인별_학습_전체_raw.xlsx',
    '개인별전체학습',
    individual_full_data,
    ['개인ID'],
    ['이름', '조직', '직책', '24년학습시간', '25년학습시간', '학습카드명', '학습일시']
)

print("\nAll Excel templates have been created successfully!")
print("Location: excel_templates/ directory")

